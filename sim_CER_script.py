""" #Import librerie  """

import pandas as pd
from Collections_script_module import OrderedDict
import subprocess  # to run the TRNSYS simulation
import time  # to measure the computation time
import pyautogui as pyautogui
import matplotlib.pyplot as plt
import numpy as np
from fractions import Fraction
import math
from openpyxl import load_workbook
import os

# import random

"""

                Codice per ottimizzazione Comunità Energetiche Rinnovabili - Mirco Tirloni

"""

"""
                                            (0) Dati per generare file excel
"""

"Calcolo area pv, impianto a terra"

# CER_peak_power = 800  # [kW_el] 4% della cabina primaria

CER_El_load_annuo = 10000000  # [kWh]

peak_power_PV_terra = 0.0000000000000001  # [kW_el]

area_pv_a_terra = peak_power_PV_terra / 0.2  # meglio usare 0,16[kW/m^2] considerando quindi una radiazione di 800[W/m^2] ed efficienza pannello di 0,2

"WT/PV ratio"

# pv_wt_ratio = [1 / 2, 1, 2, 10, 100, 200, inf]
wt_pv_ratio = [2, 1, 0.5, 0.1, 0.01, 0.02, 0.0000000000001]

n_wt_a_terra = math.ceil((peak_power_PV_terra * wt_pv_ratio[6]) / 1.5)

"se wt_pv diverso da zero settare a 1 il wt_signal per l'impianto a terra"

wt_signal = 0

" Dati tecnici edifici "

"""
0: inclinazione pannello 
1: orientamento - azimuth pannello
2: eff. pannello
3: area pannelli installati sugli edifici
4: n°turbine sugli edifici [NON Modificare]
5: capacità batteria [modificare solo insieme all'area pv]
6: COP
7: EER
8: Area Edificio
9: Volume Edificio
"""

"Superficie media utenze immobiliari "
#             RES   IND  OFF    SHO
sup_utenze = [87.2, 940, 142.9, 125.2]
"Altezza media utenze immobiliari "
h_utenze = [2.7, 6, 2.7, 2.7]
"Volume medio utenze immobiliari "
vol_utenze = [(sup_utenze[0] * h_utenze[0]),
              (sup_utenze[1] * h_utenze[1]),
              (sup_utenze[2] * h_utenze[2]),
              (sup_utenze[3] * h_utenze[3])]

"Combinazioni area pv[m^2]e capacità batterie[Wh]"
comb_pv_bat_RES = [25, 5000]
comb_pv_bat_IND = [50, 10000]
comb_pv_bat_OFF = [30, 6000]
comb_pv_bat_SHO = [25, 5000]

"Heat pump: EER,COP"

EER = 4
COP = 3.5

"Dati tecnici edifici"
techn_data_RES = [30, 0, 0.2, comb_pv_bat_RES[0], 1, comb_pv_bat_RES[1], COP, EER, sup_utenze[0], vol_utenze[0]]
techn_data_IND = [30, 0, 0.2, comb_pv_bat_IND[0], 1, comb_pv_bat_IND[1], COP, EER, sup_utenze[1], vol_utenze[1]]
techn_data_OFF = [30, 0, 0.2, comb_pv_bat_OFF[0], 1, comb_pv_bat_OFF[1], COP, EER, sup_utenze[2], vol_utenze[2]]
techn_data_SHO = [30, 0, 0.2, comb_pv_bat_SHO[0], 1, comb_pv_bat_SHO[1], COP, EER, sup_utenze[3], vol_utenze[3]]

"EL Peak power degli edifici"

El_p_pow_RES = (17.626 * techn_data_RES[9]) / 1000  # kW_el
El_p_pow_OFF = (10.171 * techn_data_OFF[9]) / 1000  # kW_el
El_p_pow_IND = (5.712 * techn_data_IND[9]) / 1000  # kW_el
El_p_pow_SHO = (20.46 * techn_data_SHO[9]) / 1000  # kW_el

"El load delle utenze [kWh]"

El_load_RES = 7074.5
El_load_OFF = 12515.73
El_load_IND = 104587.95
El_load_SHO = 30604.21

"Creazione liste consumer/prosumer con caratteristiche edificio e dati tecnici"

"                                        CONSUMER                                                                "

# lista consumer RES
lista_00 = [0, 0, 0, 0, 3, 0, 0, "nan", "nan", 0, 0, 0, 0, 0, 0, techn_data_RES[6], techn_data_RES[7],
            techn_data_RES[8],
            techn_data_RES[9]]
# lista consumer IND
lista_01 = [0, 1, 0, 0, 3, 0, 0, "nan", "nan", 0, 0, 0, 0, 0, 0, techn_data_IND[6], techn_data_IND[7],
            techn_data_IND[8],
            techn_data_IND[9]]
# lista consumer OFF
lista_02 = [0, 2, 0, 0, 3, 0, 0, "nan", "nan", 0, 0, 0, 0, 0, 0, techn_data_OFF[6], techn_data_OFF[7],
            techn_data_OFF[8],
            techn_data_OFF[9]]
# lista consumer SHO
lista_03 = [0, 3, 0, 0, 3, 0, 0, "nan", "nan", 0, 0, 0, 0, 0, 0, techn_data_SHO[6], techn_data_SHO[7],
            techn_data_SHO[8],
            techn_data_SHO[9]]

"                                        PROSUMER                                                                "

# lista prosumer RES
lista_10 = [1, 0, 1, 0, 3, 0, 1, "nan", "nan", techn_data_RES[0], techn_data_RES[1], techn_data_RES[2],
            techn_data_RES[3],
            techn_data_RES[4], techn_data_RES[5],
            techn_data_RES[6], techn_data_RES[7], techn_data_RES[8], techn_data_RES[9]]
# lista prosumer IND
lista_11 = [1, 1, 1, 0, 3, 0, 1, "nan", "nan", techn_data_IND[0], techn_data_IND[1], techn_data_IND[2],
            techn_data_IND[3],
            techn_data_IND[4], techn_data_IND[5],
            techn_data_IND[6], techn_data_IND[7], techn_data_IND[8], techn_data_IND[9]]
# lista prosumer OFF
lista_12 = [1, 2, 1, 0, 3, 0, 1, "nan", "nan", techn_data_OFF[0], techn_data_OFF[1], techn_data_OFF[2],
            techn_data_OFF[3],
            techn_data_OFF[4], techn_data_OFF[5],
            techn_data_OFF[6], techn_data_OFF[7], techn_data_OFF[8], techn_data_OFF[9]]
# lista prosumer SHO
lista_13 = [1, 3, 1, 0, 3, 0, 1, "nan", "nan", techn_data_SHO[0], techn_data_SHO[1], techn_data_SHO[2],
            techn_data_SHO[3],
            techn_data_SHO[4], techn_data_SHO[5],
            techn_data_SHO[6], techn_data_SHO[7], techn_data_SHO[8], techn_data_SHO[9]]

# lista solo rinnovabili
lista_2 = [2, "nan", 1, wt_signal, 0, 0, 0, "nan", "nan", 30, 0, 0.2, area_pv_a_terra, n_wt_a_terra, 0, 0, 0, 0, 0]

"""
"Combinazioni utenze [res, off, ind, sho] ; 31 totali"

combinazione_utenze = [
    # 0
    [1 / 4, 1 / 4, 1 / 4, 1 / 4],
    # 1, 2, 3, 4, 5, 6
    [1 / 2, 1 / 2, 0, 0], [1 / 2, 0, 1 / 2, 0], [1 / 2, 0, 0, 1 / 2],
    [0, 1 / 2, 1 / 2, 0], [0, 1 / 2, 0, 1 / 2], [0, 0, 1 / 2, 1 / 2],
    # 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18
    [1 / 4, 3 / 4, 0, 0], [1 / 4, 0, 3 / 4, 0], [1 / 4, 0, 0, 3 / 4],
    [3 / 4, 1 / 4, 0, 0], [0, 1 / 4, 3 / 4, 0], [0, 1 / 4, 0, 3 / 4],
    [3 / 4, 0, 1 / 4, 0], [0, 3 / 4, 1 / 4, 0], [0, 0, 1 / 4, 3 / 4],
    [3 / 4, 0, 0, 1 / 4], [0, 3 / 4, 0, 1 / 4], [0, 0, 3 / 4, 1 / 4],
    # 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30
    [1 / 2, 1 / 4, 1 / 4, 0], [1 / 2, 0, 1 / 4, 1 / 4], [1 / 2, 1 / 4, 0, 1 / 4],
    [1 / 4, 1 / 2, 1 / 4, 0], [1 / 4, 1 / 2, 0, 1 / 4], [0, 1 / 2, 1 / 4, 1 / 4],
    [0, 1 / 4, 1 / 2, 1 / 4], [1 / 4, 0, 1 / 2, 1 / 4], [1 / 4, 1 / 4, 1 / 2, 0],
    [0, 1 / 4, 1 / 4, 1 / 2], [1 / 4, 0, 1 / 4, 1 / 2], [1 / 4, 1 / 4, 0, 1 / 2]
]

# pros_cons_ratio = [Fraction(1, 2), Fraction(1, 3), Fraction(1, 4), Fraction(1, 5), Fraction(1, 6),Fraction(1, 7), Fraction(1, 8), Fraction(1, 9), Fraction(1, 10), Fraction(1, 15)]
"""

"""
                                Creazione liste backup
"""

# Creazione liste backup per edifici consumer
# RES
E_abs_consumers_RES_b = np.zeros(8760)
# OFF
E_abs_consumers_OFF_b = np.zeros(8760)
# IND
E_abs_consumers_IND_b = np.zeros(8760)
# SHO
E_abs_consumers_SHO_b = np.zeros(8760)

# Creazione liste backup per edifici prosumer
# RES
Delta_E_pros_RES_b = np.zeros(8760)
Autocons_pros_RES_b = np.zeros(8760)
list_El_load_pros_RES_b = np.zeros(8760)
list_El_gen_pros_RES_b = np.zeros(8760)
# OFF
Delta_E_pros_OFF_b = np.zeros(8760)
Autocons_pros_OFF_b = np.zeros(8760)
list_El_load_pros_OFF_b = np.zeros(8760)
list_El_gen_pros_OFF_b = np.zeros(8760)
# IND
Delta_E_pros_IND_b = np.zeros(8760)
Autocons_pros_IND_b = np.zeros(8760)
list_El_load_pros_IND_b = np.zeros(8760)
list_El_gen_pros_IND_b = np.zeros(8760)
# SHO
Delta_E_pros_SHO_b = np.zeros(8760)
Autocons_pros_SHO_b = np.zeros(8760)
list_El_load_pros_SHO_b = np.zeros(8760)
list_El_gen_pros_SHO_b = np.zeros(8760)

Liste_backup_cons = [E_abs_consumers_RES_b,
                     E_abs_consumers_OFF_b,
                     E_abs_consumers_IND_b,
                     E_abs_consumers_SHO_b]

Liste_backup_pros = [[Delta_E_pros_RES_b, Autocons_pros_RES_b, list_El_load_pros_RES_b, list_El_gen_pros_RES_b],
                     [Delta_E_pros_OFF_b, Autocons_pros_OFF_b, list_El_load_pros_OFF_b, list_El_gen_pros_OFF_b],
                     [Delta_E_pros_IND_b, Autocons_pros_IND_b, list_El_load_pros_IND_b, list_El_gen_pros_IND_b],
                     [Delta_E_pros_SHO_b, Autocons_pros_SHO_b, list_El_load_pros_SHO_b, list_El_gen_pros_SHO_b]]

Lista_backup_producer = np.zeros(8760)


def excel_gen(num_car, num_dati, excel_data_path, denom_p_c_ratio, ratio_RES, ratio_OFF, ratio_IND,
              ratio_SHO):
    """Def Prosumer/Consumer ratio"""

    pros_cons_ratio = Fraction(1, denom_p_c_ratio)

    comb_utenze = [ratio_RES, ratio_OFF, ratio_IND, ratio_SHO]

    """Calcolo numero edifici"""
    lista_save_ris = []
    A = np.array([[El_load_RES, El_load_OFF, El_load_IND, El_load_SHO, 0],
                  [1, 0, 0, 0, -ratio_RES],
                  [0, 1, 0, 0, -ratio_OFF],
                  [0, 0, 1, 0, -ratio_IND],
                  [0, 0, 0, 1, -ratio_SHO],
                  ])
    b = np.array([CER_El_load_annuo, 0, 0, 0, 0])

    # Solve the linear system Ax = b
    x = np.linalg.solve(A, b)

    # numeri arrotondati per difetto

    n_ed_RES = math.floor(x[0])
    n_ed_OFF = math.floor(x[1])
    n_ed_IND = math.floor(x[2])
    n_ed_SHO = math.floor(x[3])
    n_ed_tot = math.floor(x[4])

    somma_edifici = n_ed_RES + n_ed_IND + n_ed_OFF + n_ed_SHO

    somma_peak_power = n_ed_RES * El_p_pow_RES + n_ed_IND * El_p_pow_IND + n_ed_OFF * El_p_pow_OFF + n_ed_SHO * El_p_pow_SHO
    # somma_El_load = n_ed_RES * El_load_RES + n_ed_OFF * El_load_OFF + n_ed_IND * El_load_IND + n_ed_SHO * El_load_SHO
    # se arrotondando per difetto il numero degli edifici non torna il totale:
    # aumento edifici residenziali

    if somma_edifici < n_ed_tot:
        y = n_ed_tot - somma_edifici
        n_ed_RES = n_ed_RES + y

    "calcolo n° prosumer (arrotondati per eccesso) e consumer per ogni tipologia d'utenza"

    n_pros_RES = math.ceil(n_ed_RES / (pros_cons_ratio.numerator + pros_cons_ratio.denominator))
    n_cons_RES = n_ed_RES - n_pros_RES

    n_pros_OFF = math.ceil(n_ed_OFF / (pros_cons_ratio.numerator + pros_cons_ratio.denominator))
    n_cons_OFF = n_ed_OFF - n_pros_OFF

    n_pros_IND = math.ceil(n_ed_IND / (pros_cons_ratio.numerator + pros_cons_ratio.denominator))
    n_cons_IND = n_ed_IND - n_pros_IND

    n_pros_SHO = math.ceil(n_ed_SHO / (pros_cons_ratio.numerator + pros_cons_ratio.denominator))
    n_cons_SHO = n_ed_SHO - n_pros_SHO

    "Calcolo intervalli excel lines"

    # n_pros_RES
    inf_pros_RES = 8
    sup_pros_RES = 8 + n_pros_RES

    # n_cons_RES
    inf_cons_RES = sup_pros_RES
    sup_cons_RES = inf_cons_RES + n_cons_RES

    # n_pros_OFF
    inf_pros_OFF = sup_cons_RES
    sup_pros_OFF = inf_pros_OFF + n_pros_OFF

    # n_cons_OFF
    inf_cons_OFF = sup_pros_OFF
    sup_cons_OFF = inf_cons_OFF + n_cons_OFF

    # n_pros_IND
    inf_pros_IND = sup_cons_OFF
    sup_pros_IND = inf_pros_IND + n_pros_IND

    # n_cons_IND
    inf_cons_IND = sup_pros_IND
    sup_cons_IND = inf_cons_IND + n_cons_IND

    # n_pros_SHO
    inf_pros_SHO = sup_cons_IND
    sup_pros_SHO = inf_pros_SHO + n_pros_SHO

    # n_cons_SHO
    inf_cons_SHO = sup_pros_SHO
    sup_cons_SHO = inf_cons_SHO + n_cons_SHO

    "Calcolo colonne da sovrascrivere"

    n_columns = 4 + num_car + num_dati  # in base al file excel TEMPLATE

    excel_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                     'U', 'V', 'W', 'X', 'Y', 'Z']

    # Get the absolute path of the file
    file_path_excel = os.path.abspath(excel_data_path)

    # Load an existing workbook or create a new one
    workbook = load_workbook(file_path_excel)

    # Select the active worksheet

    sheet = workbook["Scheda_dati_CER"]

    "pulizia vecchio file"
    sheet.delete_rows(7, sheet.max_row)

    # Modify cells
    numero_edificio = 1  # inizializzazione variabile

    "Sovrascrivo Impianto rinnovabile a terra "

    sheet[excel_columns[1] + str(7)] = "Edificio_" + str(numero_edificio)
    j = 0

    for n in range(2, n_columns):
        sheet[excel_columns[n] + str(7)] = lista_2[j]
        j = j + 1
        #
    numero_edificio = numero_edificio + 1

    "Sovrascrivo Edifici RES"

    if n_ed_RES != 0:

        "Sovrascrivo Edifici PROS - RES"

        for i in range(inf_pros_RES, sup_pros_RES):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_10[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #
        "Sovrascrivo Edifici CONS - RES"

        for i in range(inf_cons_RES, sup_cons_RES):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_00[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #

    #

    "Sovrascrivo Edifici OFF"

    if n_ed_OFF != 0:

        "Sovrascrivo Edifici PROS - OFF"

        for i in range(inf_pros_OFF, sup_pros_OFF):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_12[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #
        "Sovrascrivo Edifici CONS - OFF"

        for i in range(inf_cons_OFF, sup_cons_OFF):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_02[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #

    #
    "Sovrascrivo Edifici IND"

    if n_ed_IND != 0:

        "Sovrascrivo Edifici PROS - IND"

        for i in range(inf_pros_IND, sup_pros_IND):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_11[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #
        "Sovrascrivo Edifici CONS - IND"

        for i in range(inf_cons_IND, sup_cons_IND):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_01[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #

    #
    "Sovrascrivo Edifici SHO"

    if n_ed_SHO != 0:

        "Sovrascrivo Edifici PROS - SHO"

        for i in range(inf_pros_SHO, sup_pros_SHO):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_13[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #
        "Sovrascrivo Edifici CONS - SHO"

        for i in range(inf_cons_SHO, sup_cons_SHO):

            sheet[excel_columns[1] + str(i)] = "Edificio_" + str(numero_edificio)

            j = 0

            for n in range(2, n_columns):
                sheet[excel_columns[n] + str(i)] = lista_03[j]
                j = j + 1
                #

            numero_edificio = numero_edificio + 1
        #

    #

    # Save the changes to the workbook
    workbook.save(file_path_excel)

    # lista per salvare risultati su excel

    lista_save_ris.append(str(comb_utenze))
    lista_save_ris.append(str(pros_cons_ratio))
    lista_save_ris.append(n_ed_tot)
    # lista_save_ris.append(somma_El_load)
    lista_save_ris.append(somma_peak_power)
    lista_save_ris.append(peak_power_PV_terra)
    lista_save_ris.append(wt_pv_ratio[6])

    return lista_save_ris, n_ed_tot, somma_peak_power
    #


"""
                                            (1) Raccolta dati 
"""
"""
                                        (1.1) Raccolta dati CER
"""

"   #Import da excel "


def collect_data(excl_data_path, n_car, n_dati, txt_var, n_ed_tot):
    read_data = pd.read_excel(excl_data_path)

    n_ed = n_ed_tot + 1  # +1 perchè non viene considerato l'impianto a terra nel calcolo di n_ed_tot"

    "   #Lista Edifici  "

    lista_edifici = []

    for i in range(5, n_ed + + 5):  # modificare l'ultima riga da leggere
        lista_edifici.append(read_data.iat[i, 1])
    print("")
    print("numero edifici:", n_ed)
    txt_var.append(["numero edifici:", n_ed])
    # print(lista_edifici)

    "   #Assegnazione caratteristiche edifici   "

    #     Legenda:

    # 1   prosumer/consumer       :   0: consumer     1: prosumer
    # 2   Tipologia di edificio   :   0: residenziale 1: industriale         2: ufficio/negozio
    # 3   P.V.                    :   0: no           1: si
    # 4   W.T.                    :   0: no           1: si
    # 5   Heat Pump               :   0: no                     1: solo riscaldamento
    #                                 2: solo condizionamento   3: pompa reversibile
    # 6   Auto elettrica          :   0: no           1: si
    # 7   Condizionamento estivo  :   0: no           1: si
    # 8   Accumulo con batterie   :   0: no           1: si

    carat_edif = []
    carat_edif_ordered = []

    for i in range(0, n_ed):
        carat_edif.append({})
        for x in range(0, n_car):
            A = read_data.iat[0, x + 2]
            B = read_data.iat[i + 5, x + 2]
            carat_edif[i][A] = float(B)
            #
        carat_edif_ordered.append([])
        carat_edif_ordered[i] = OrderedDict(carat_edif[i])
        #
    #

    " Assegnazione codice agli edifici consumer/prosumer"

    codici_edifici_prosumer = []
    codici_edifici_consumer = []
    codici_solo_rinnovabili = []

    for i in range(0, n_ed):
        if list(carat_edif_ordered[i].values())[0] == 0:
            codici_edifici_consumer.append(i)
        elif list(carat_edif_ordered[i].values())[0] == 1:
            codici_edifici_prosumer.append(i)
        else:
            codici_solo_rinnovabili.append(i)

    txt_var.append(["n°prosumer:", len(codici_edifici_prosumer)])
    txt_var.append(["n°consumer:", len(codici_edifici_consumer)])
    txt_var.append(["n°solo rinn.:", len(codici_solo_rinnovabili)])

    " Assegnazione codice agli edifici RES/IND/OFF"

    cod_ed_RES = []
    cod_ed_IND = []
    cod_ed_OFF = []
    cod_ed_SHO = []

    for i in range(0, n_ed):
        if list(carat_edif_ordered[i].values())[1] == 0:
            cod_ed_RES.append(i)
        elif list(carat_edif_ordered[i].values())[1] == 1:
            cod_ed_IND.append(i)
        elif list(carat_edif_ordered[i].values())[1] == 2:
            cod_ed_OFF.append(i)
        elif list(carat_edif_ordered[i].values())[1] == 3:
            cod_ed_SHO.append(i)

    "   #Assegnazione scheda tecnica edifici    "

    #    Legenda:

    # 1   Dettagli edificio   :   Area,   Volume
    # 2   P.V.                :   Area,   Inclinazione,   Orientamento,   Potenza
    # 3   W.T.                :   Potenza
    # 4   Auto Elettrica      :   Ricarica alle 12.00(1=si,0=no), Ricarica alle 24.00(1=si,0=no), Potenza assorbita
    # 5   Accumulo            :   Capacità batteria
    # 6   Climatizzazione     :   N°HP installate,   cooling load,   heating load,   cooling mode,   heating mode

    # n_dati = 15  # numero dati tecnici importati
    scheda_tecnica = []  # lista di dizionari contenenti i dati tecnici di ogni edificio
    scheda_tecnica_ordered = []  # dizionari con elementi "numerati"
    area_PV = 0  # Area PV installata [m^2]
    for i in range(0, n_ed):
        scheda_tecnica.append({})
        for x in range(0, n_dati):
            A = read_data.iat[2, x + 11]  # 11 = colonna da cui si inizia ad importare i dati
            B = read_data.iat[i + 5, x + 11]
            scheda_tecnica[i][A] = float(B)
            #
        scheda_tecnica_ordered.append([])
        scheda_tecnica_ordered[i] = OrderedDict(scheda_tecnica[i])
        #
        area_PV = area_PV + (list(scheda_tecnica_ordered[i].values()))[3]

    #

    return n_ed, codici_edifici_consumer, codici_edifici_prosumer, codici_solo_rinnovabili, lista_edifici, scheda_tecnica_ordered, carat_edif_ordered, cod_ed_RES, cod_ed_IND, cod_ed_OFF, cod_ed_SHO, area_PV


"""
                                        (1.2) Raccolta dati PUN
"""


def collect_PUN():
    excl_data_path_PUN = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\PUN_dati.xlsx"
    read_data = pd.read_excel(excl_data_path_PUN)
    PUN_list = []
    for i in range(0, 8760):
        PUN_list.append(read_data.iat[i, 0])

    return PUN_list


"""
                                    Vettore sim.hours
"""
hours = []
for a in range(1, 8761):
    hours.append(a)

"""
                                    (2) def. function
"""

"""                                    
                                (2.1)   Read txt odd lines                                
"""


# Definizione della funzione per leggere le righe dispari e aggiungerle in una slice
# DISPARI PERCHE' LA PRIMA RIGA E' LA RIGA NUMERO 0.

def leggi_righe_dispari(file_path):
    slice_righe_dispari = []  # Inizializzazione della slice vuota

    with open(file_path, 'r') as file:
        lines = file.readlines()
        for i in range(3, len(lines), 2):  # Itera sulle righe dispari saltando la simulazione all'ora 0:00
            slice_righe_dispari.append(float(lines[i].strip()))  # Aggiunge la riga senza spazi bianchi alla slice

    return slice_righe_dispari


"""                                    
                                (2.2)   Modifica file dck                                
"""


def edit_dck(n_dati, dck_temp_file_path, dck_file_path, cod_ed, cod_ed_RES, cod_ed_IND, cod_ed_OFF, cod_ed_SHO,
             scheda_tecnica_ordered, carat_edif_ordered):
    """   #Opening dublicated template .dck file """

    with open(
            dck_temp_file_path,
            'r') as file_in:
        filedata = file_in.read()

    "   #Changing/replacing the pyt tags to parameter values in the .dck text   "

    # template variables

    py_t_var = ['py_tv_1', 'py_tv_2', 'py_tv_3', 'py_tv_4', 'py_tv_5', 'py_tv_6', 'py_tv_7', 'py_tv_8', 'py_tv_9',
                'py_tv_A']

    # Replacing dati tecnici

    for i in range(0, n_dati):
        filedata = filedata.replace(py_t_var[i], str(list(scheda_tecnica_ordered[cod_ed].values())[i]))
        # inserire la riga di excel che si vuole sovrascrivere nel template.dck
        # NB: scheda_tecnica_ordered[n] contiene i dati tecnici da sovrascrivere nel dck
        # n = elemento n-esimo della lista scheda_tecnica_ordered
        # n=0 : Edificio_1, n=1 : Edificio_2, ecc...

    RES_LOADS = ['RES_CO_LOAD.txt', 'RES_TH_LOAD.txt', 'RES_EL_LOAD.txt']
    IND_LOADS = ['IND_CO_LOAD.txt', 'IND_TH_LOAD.txt', 'IND_EL_LOAD.txt']
    OFF_LOADS = ['OFF_CO_LOAD.txt', 'OFF_TH_LOAD.txt', 'OFF_EL_LOAD.txt']
    SHO_LOADS = ['SHO_CO_LOAD.txt', 'SHO_TH_LOAD.txt', 'SHO_EL_LOAD.txt']

    # Replacing user loads (EL,TH,CO)

    if cod_ed in cod_ed_RES:
        filedata = filedata.replace('py_tv_X', RES_LOADS[0])
        filedata = filedata.replace('py_tv_Y', RES_LOADS[1])
        filedata = filedata.replace('py_tv_Z', RES_LOADS[2])
    elif cod_ed in cod_ed_IND:
        filedata = filedata.replace('py_tv_X', IND_LOADS[0])
        filedata = filedata.replace('py_tv_Y', IND_LOADS[1])
        filedata = filedata.replace('py_tv_Z', IND_LOADS[2])
    elif cod_ed in cod_ed_OFF:
        filedata = filedata.replace('py_tv_X', OFF_LOADS[0])
        filedata = filedata.replace('py_tv_Y', OFF_LOADS[1])
        filedata = filedata.replace('py_tv_Z', OFF_LOADS[2])
    elif cod_ed in cod_ed_SHO:
        filedata = filedata.replace('py_tv_X', SHO_LOADS[0])
        filedata = filedata.replace('py_tv_Y', SHO_LOADS[1])
        filedata = filedata.replace('py_tv_Z', SHO_LOADS[2])

    # Replacing wt control signal

    filedata = filedata.replace('py_tv_J', str(list(carat_edif_ordered[cod_ed].values())[3]))

    with open(
            dck_file_path,
            'w') as dckfile_out:
        dckfile_out.write(filedata)


"""                                    
                                (2.3)   Simulazione edificio con sim studio                                     
"""


def run_sim_studio(dck_file_path, codice_ed, carat_edif_ordered):
    """   #Open TRNSYS project """

    subprocess.Popen(['start', '',
                      r"C:\TRNSYS18\Studio\Exe\Studio.exe"],
                     shell=True)

    "#Import dck file "
    #
    time.sleep(3)  # tempo per aprire trnsys
    pyautogui.press('alt')
    time.sleep(0.05)

    x = 0
    while x < 3:  # arrow down 3 times
        pyautogui.press('down')
        time.sleep(0.05)
        x = x + 1

    pyautogui.press('enter')
    time.sleep(1.0)  # tempo per aprire la finestra
    pyautogui.typewrite(dck_file_path)
    time.sleep(0.05)
    pyautogui.press('enter')
    time.sleep(2)  # tempo per importare il file dck nuovo

    "#Running simulation "
    #
    pyautogui.press('f8')
    #
    if list(carat_edif_ordered[codice_ed].values())[0] == 0:
        time.sleep(5)  # tempo per simulare ed consumer
    elif list(carat_edif_ordered[codice_ed].values())[0] == 1:
        time.sleep(21)  # tempo per simulare ed prosumer
    else:
        time.sleep(8)  # tempo per simulare solo rinnovabili

    "#Exit online plotter"
    #
    pyautogui.press('enter')
    time.sleep(0.05)

    "#Exit trnsys"
    #
    pyautogui.hotkey('alt', 'f4')
    time.sleep(0.05)


"""                                    
                                (2.4)   Lettura e salvataggio risultati di output                          
"""


def read_txt(txt_file_path, saving_list, sav_list_ind, txt_var):
    """   #Lettura file output    """
    txt_data = []

    try:
        txt_data = leggi_righe_dispari(txt_file_path)
        # print("Mean_energy abs. or prod. = ",
        #     (sum(txt_data) / len(txt_data)))  # media aritmetica annuale dell'energia prodotta/assorbita
        # txt_var.append(["Mean_energy abs. or prod. = ",
        # (sum(txt_data) / len(txt_data))])
    except FileNotFoundError:
        print("Il file specificato non è stato trovato.")

    except Exception as e:
        print("Si è verificato un errore:", e)

    for x in range(0, len(txt_data)):
        saving_list[sav_list_ind][x] = txt_data[x]


"""
                                (2.5)   Eliminazione file di sim studio "imported" dalla directory
"""


def delete_sim_file(directory_del_files, files_to_delete):
    # Specify the directory containing the files you want to delete

    # Get a list of files to delete

    # Delete the filtered files

    for filename in files_to_delete:
        file_path = os.path.join(directory_del_files, filename)
        try:
            os.remove(file_path)
            # print(f"File '{file_path}' deleted successfully.")
        except OSError as e:
            print(f"Error: {e.filename} - {e.strerror}.")


"""
                                (2.6)   Simulazione completa singolo edificio
"""


def simulation(n_dati, dck_temp_file_path, dck_file_path, cod_ed, sav_list_ind, cod_ed_RES, cod_ed_IND, cod_ed_OFF,
               cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered, txt_file_path, saving_list, directory_del_files,
               files_to_delete, txt_var):
    try:
        #   modifica file dck template
        edit_dck(n_dati, dck_temp_file_path, dck_file_path, cod_ed, cod_ed_RES, cod_ed_IND, cod_ed_OFF, cod_ed_SHO,
                 scheda_tecnica_ordered, carat_edif_ordered)
        # time.sleep(0.5)
        #   simulazione con sim studio
        run_sim_studio(dck_file_path, cod_ed, carat_edif_ordered)
        # time.sleep(0.5)
        #   lettura e salvataggio output txt file
        read_txt(txt_file_path, saving_list, sav_list_ind, txt_var)
        # time.sleep(0.5)
        #   pulizia dati simulazione sim studio
        delete_sim_file(directory_del_files, files_to_delete)
        # time.sleep(0.5)
        #

    except FileNotFoundError:
        print("Il file specificato non è stato trovato.")


"""
                                (3)  Input simulazioni
"""

"""
                                (3.1)  Input consumers
"""

# E_abs_consumers = [0] * 8760

# dati input

dck_temp_file_path_A = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Edifici Consumer\Consumer_Model_TEMPLATE.dck"

dck_file_path_A = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Edifici Consumer\Consumer_Model.dck"

txt_file_path_A = "C:\\Users\\Mirco Tirloni\\Desktop\\TESI_Mirco\\Codice_TESI_Mirco\\Trnsys files TESI_Mirco\\Edifici Consumer\\energy_plot.txt"

directory_del_files_A = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Edifici Consumer"

files_to_delete_A = ['Consumer_Model_imported.dck', 'Consumer_Model_imported.log', 'Consumer_Model_imported.lst',
                     'Consumer_Model_imported.tpf']

"""
                                (3.2)  Input prosumers
"""

# dati input

dck_temp_file_path_B = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Edifici Prosumer\Prosumer_Model_TEMPLATE.dck"

dck_file_path_B = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Edifici Prosumer\Prosumer_Model.dck"

txt_file_path_B_1 = "C:\\Users\\Mirco Tirloni\\Desktop\\TESI_Mirco\\Codice_TESI_Mirco\\Trnsys files TESI_Mirco\\Edifici Prosumer\\energy_plot.txt"

txt_file_path_B_2 = "C:\\Users\\Mirco Tirloni\\Desktop\\TESI_Mirco\\Codice_TESI_Mirco\\Trnsys files TESI_Mirco\\Edifici Prosumer\\autoconsumo.txt"

txt_file_path_B_3 = "C:\\Users\\Mirco Tirloni\\Desktop\\TESI_Mirco\\Codice_TESI_Mirco\\Trnsys files TESI_Mirco\\Edifici Prosumer\\E_load_pros.txt"

txt_file_path_B_4 = "C:\\Users\\Mirco Tirloni\\Desktop\\TESI_Mirco\\Codice_TESI_Mirco\\Trnsys files TESI_Mirco\\Edifici Prosumer\\E_gen_pros.txt"

directory_del_files_B = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Edifici Prosumer"

files_to_delete_B = ['Prosumer_Model_imported.tpf', 'Prosumer_Model_imported.dck', 'Prosumer_Model_imported.log',
                     'Prosumer_Model_imported.lst']

"""
                                (3.3)  Input solo rinnovabili
"""

# dati input

dck_temp_file_path_C = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Solo Rinnovabili\Solo_Rinnovabili_TEMPLATE.dck"

dck_file_path_C = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Solo Rinnovabili\Solo_Rinnovabili.dck"

txt_file_path_C = "C:\\Users\\Mirco Tirloni\\Desktop\\TESI_Mirco\\Codice_TESI_Mirco\\Trnsys files TESI_Mirco\\Solo Rinnovabili\\energy_plot.txt"

directory_del_files_C = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Trnsys files TESI_Mirco\Solo Rinnovabili"

files_to_delete_C = ['Solo_Rinnovabili_imported.tpf', 'Solo_Rinnovabili_imported.dck',
                     'Solo_Rinnovabili_imported.log', 'Solo_Rinnovabili_imported.lst']

"""
                              (4) def. fun. per simulazione CER
"""

"""
                              (4.1) def. simulazione edifici CER
"""


def sim_edifici(n_dati, n_ed, codici_edifici_consumer, codici_edifici_prosumer, codici_solo_rinnovabili,
                lista_edifici,
                cod_ed_RES, cod_ed_IND, cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered, txt_var,
                Backup_cons, Backup_pros, Backup_producer):
    # azzeramento vettori (saving list)
    """
    Delta_E_pros = [[0] * 8760 for _ in range(len(codici_edifici_prosumer))]
    Autocons_pros = [[0] * 8760 for _ in range(len(codici_edifici_prosumer))]
    list_El_load_pros = [[0] * 8760 for _ in range(len(codici_edifici_prosumer))]
    list_El_gen_pros = [[0] * 8760 for _ in range(len(codici_edifici_prosumer))]
    Gen_rin = [[0] * 8760 for _ in range(len(codici_solo_rinnovabili))]
    E_abs_consumers = [[0] * 8760 for _ in range(len(codici_edifici_consumer))]
    """
    Delta_E_pros = np.zeros((len(codici_edifici_prosumer), 8760))
    Autocons_pros = np.zeros((len(codici_edifici_prosumer), 8760))
    list_El_load_pros = np.zeros((len(codici_edifici_prosumer), 8760))
    list_El_gen_pros = np.zeros((len(codici_edifici_prosumer), 8760))
    Gen_rin = np.zeros((len(codici_solo_rinnovabili), 8760))
    E_abs_consumers = np.zeros((len(codici_edifici_consumer), 8760))

    # somma_liste= sum(Liste_backup_cons)+sum(Liste_backup_pros)+sum(Lista_backup_producer)

    for i in range(0, n_ed):  # ciclo sui cod.ed

        # controllo se l'edificio attuale è uguale a quello di prima e nel caso copio i suoi dati

        if i != 0 and list(scheda_tecnica_ordered[i].values()) == list(scheda_tecnica_ordered[i - 1].values()) and list(
                carat_edif_ordered[i].values()) == list(carat_edif_ordered[i - 1].values()):

            if i in codici_edifici_consumer:
                x = codici_edifici_consumer.index(i)
                np.copyto(E_abs_consumers[x], E_abs_consumers[x - 1])
            elif i in codici_edifici_prosumer:
                x = codici_edifici_prosumer.index(i)
                np.copyto(Delta_E_pros[x], Delta_E_pros[x - 1])
                np.copyto(Autocons_pros[x], Autocons_pros[x - 1])
                np.copyto(list_El_load_pros[x], list_El_load_pros[x - 1])
                np.copyto(list_El_gen_pros[x], list_El_gen_pros[x - 1])

            elif i in codici_solo_rinnovabili:
                x = codici_edifici_prosumer.index(i)
                np.copyto(Gen_rin[x], Gen_rin[x - 1])

        else:

            if i in codici_edifici_consumer:  # controllo se il cod.ed è nella lista consumer

                if i in cod_ed_RES:  # controllo se il cod.ed è nella lista RES

                    if np.all(Backup_cons[0] == 0):  # controllo se è presente un backup

                        sav_list_ind = codici_edifici_consumer.index(i)
                        # tramite codice edificio risalgo alla posizione nella lista dei consumer
                        # registro la posizione come sav_list_index
                        print("Simulazione:", lista_edifici[i], ", consumer", sav_list_ind + 1, "/",
                              len(codici_edifici_consumer))
                        simulation(n_dati, dck_temp_file_path_A, dck_file_path_A, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered, txt_file_path_A,
                                   E_abs_consumers, directory_del_files_A, files_to_delete_A, txt_var)
                        # copia risultati nella lista backup
                        np.copyto(Backup_cons[0], E_abs_consumers[codici_edifici_consumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(E_abs_consumers[codici_edifici_consumer.index(i)], Backup_cons[0])
                elif i in cod_ed_OFF:

                    if np.all(Backup_cons[1] == 0):

                        sav_list_ind = codici_edifici_consumer.index(i)
                        # tramite codice edificio risalgo alla posizione nella lista dei consumer
                        # registro la posizione come sav_list_index
                        print("Simulazione:", lista_edifici[i], ", consumer", sav_list_ind + 1, "/",
                              len(codici_edifici_consumer))
                        simulation(n_dati, dck_temp_file_path_A, dck_file_path_A, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered, txt_file_path_A,
                                   E_abs_consumers, directory_del_files_A, files_to_delete_A, txt_var)
                        # copia risultati nella lista backup
                        np.copyto(Backup_cons[1], E_abs_consumers[codici_edifici_consumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(E_abs_consumers[codici_edifici_consumer.index(i)], Backup_cons[1])
                elif i in cod_ed_IND:

                    if np.all(Backup_cons[2] == 0):

                        sav_list_ind = codici_edifici_consumer.index(i)
                        # tramite codice edificio risalgo alla posizione nella lista dei consumer
                        # registro la posizione come sav_list_index
                        print("Simulazione:", lista_edifici[i], ", consumer", sav_list_ind + 1, "/",
                              len(codici_edifici_consumer))

                        simulation(n_dati, dck_temp_file_path_A, dck_file_path_A, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered, txt_file_path_A,
                                   E_abs_consumers, directory_del_files_A, files_to_delete_A, txt_var)
                        # copia risultati nella lista backup
                        np.copyto(Backup_cons[2], E_abs_consumers[codici_edifici_consumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(E_abs_consumers[codici_edifici_consumer.index(i)], Backup_cons[2])

                elif i in cod_ed_SHO:

                    if np.all(Backup_cons[3] == 0):

                        sav_list_ind = codici_edifici_consumer.index(i)
                        # tramite codice edificio risalgo alla posizione nella lista dei consumer
                        # registro la posizione come sav_list_index
                        print("Simulazione:", lista_edifici[i], ", consumer", sav_list_ind + 1, "/",
                              len(codici_edifici_consumer))

                        simulation(n_dati, dck_temp_file_path_A, dck_file_path_A, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered, txt_file_path_A,
                                   E_abs_consumers, directory_del_files_A, files_to_delete_A, txt_var)
                        # copia risultati nella lista backup
                        np.copyto(Backup_cons[3], E_abs_consumers[codici_edifici_consumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(E_abs_consumers[codici_edifici_consumer.index(i)], Backup_cons[3])

            elif i in codici_edifici_prosumer:

                if i in cod_ed_RES:

                    if np.all(Backup_pros[0][1] == 0):

                        sav_list_ind = codici_edifici_prosumer.index(i)

                        print("Simulazione:", lista_edifici[i], ", prosumer", sav_list_ind + 1, "/",
                              len(codici_edifici_prosumer))

                        simulation(n_dati, dck_temp_file_path_B, dck_file_path_B, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered,
                                   txt_file_path_B_1,
                                   Delta_E_pros, directory_del_files_B, files_to_delete_B, txt_var)

                        read_txt(txt_file_path_B_2, Autocons_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_3, list_El_load_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_4, list_El_gen_pros, sav_list_ind, txt_var)

                        # copia risultati nella lista backup
                        np.copyto(Backup_pros[0][0], Delta_E_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[0][1], Autocons_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[0][2], list_El_load_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[0][3], list_El_gen_pros[codici_edifici_prosumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(Delta_E_pros[codici_edifici_prosumer.index(i)], Backup_pros[0][0])
                        np.copyto(Autocons_pros[codici_edifici_prosumer.index(i)], Backup_pros[0][1])
                        np.copyto(list_El_load_pros[codici_edifici_prosumer.index(i)], Backup_pros[0][2])
                        np.copyto(list_El_gen_pros[codici_edifici_prosumer.index(i)], Backup_pros[0][3])

                elif i in cod_ed_OFF:

                    if np.all(Backup_pros[1][1] == 0):

                        sav_list_ind = codici_edifici_prosumer.index(i)

                        print("Simulazione:", lista_edifici[i], ", prosumer", sav_list_ind + 1, "/",
                              len(codici_edifici_prosumer))

                        simulation(n_dati, dck_temp_file_path_B, dck_file_path_B, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered,
                                   txt_file_path_B_1,
                                   Delta_E_pros, directory_del_files_B, files_to_delete_B, txt_var)

                        read_txt(txt_file_path_B_2, Autocons_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_3, list_El_load_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_4, list_El_gen_pros, sav_list_ind, txt_var)

                        # copia risultati nella lista backup
                        np.copyto(Backup_pros[1][0], Delta_E_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[1][1], Autocons_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[1][2], list_El_load_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[1][3], list_El_gen_pros[codici_edifici_prosumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(Delta_E_pros[codici_edifici_prosumer.index(i)], Backup_pros[1][0])
                        np.copyto(Autocons_pros[codici_edifici_prosumer.index(i)], Backup_pros[1][1])
                        np.copyto(list_El_load_pros[codici_edifici_prosumer.index(i)], Backup_pros[1][2])
                        np.copyto(list_El_gen_pros[codici_edifici_prosumer.index(i)], Backup_pros[1][3])

                elif i in cod_ed_IND:

                    if np.all(Backup_pros[2][1] == 0):

                        sav_list_ind = codici_edifici_prosumer.index(i)

                        print("Simulazione:", lista_edifici[i], ", prosumer", sav_list_ind + 1, "/",
                              len(codici_edifici_prosumer))

                        simulation(n_dati, dck_temp_file_path_B, dck_file_path_B, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered,
                                   txt_file_path_B_1,
                                   Delta_E_pros, directory_del_files_B, files_to_delete_B, txt_var)

                        read_txt(txt_file_path_B_2, Autocons_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_3, list_El_load_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_4, list_El_gen_pros, sav_list_ind, txt_var)

                        # copia risultati nella lista backup
                        np.copyto(Backup_pros[2][0], Delta_E_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[2][1], Autocons_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[2][2], list_El_load_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[2][3], list_El_gen_pros[codici_edifici_prosumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(Delta_E_pros[codici_edifici_prosumer.index(i)], Backup_pros[2][0])
                        np.copyto(Autocons_pros[codici_edifici_prosumer.index(i)], Backup_pros[2][1])
                        np.copyto(list_El_load_pros[codici_edifici_prosumer.index(i)], Backup_pros[2][2])
                        np.copyto(list_El_gen_pros[codici_edifici_prosumer.index(i)], Backup_pros[2][3])

                elif i in cod_ed_SHO:

                    if np.all(Backup_pros[3][1] == 0):

                        sav_list_ind = codici_edifici_prosumer.index(i)

                        print("Simulazione:", lista_edifici[i], ", prosumer", sav_list_ind + 1, "/",
                              len(codici_edifici_prosumer))

                        simulation(n_dati, dck_temp_file_path_B, dck_file_path_B, i, sav_list_ind, cod_ed_RES,
                                   cod_ed_IND,
                                   cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered,
                                   txt_file_path_B_1,
                                   Delta_E_pros, directory_del_files_B, files_to_delete_B, txt_var)

                        read_txt(txt_file_path_B_2, Autocons_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_3, list_El_load_pros, sav_list_ind, txt_var)
                        read_txt(txt_file_path_B_4, list_El_gen_pros, sav_list_ind, txt_var)

                        # copia risultati nella lista backup
                        np.copyto(Backup_pros[3][0], Delta_E_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[3][1], Autocons_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[3][2], list_El_load_pros[codici_edifici_prosumer.index(i)])
                        np.copyto(Backup_pros[3][3], list_El_gen_pros[codici_edifici_prosumer.index(i)])
                    else:
                        # prendo i risultati dalla lista backup
                        np.copyto(Delta_E_pros[codici_edifici_prosumer.index(i)], Backup_pros[3][0])
                        np.copyto(Autocons_pros[codici_edifici_prosumer.index(i)], Backup_pros[3][1])
                        np.copyto(list_El_load_pros[codici_edifici_prosumer.index(i)], Backup_pros[3][2])
                        np.copyto(list_El_gen_pros[codici_edifici_prosumer.index(i)], Backup_pros[3][3])

            elif i in codici_solo_rinnovabili:

                if np.all(Backup_producer == 0):

                    sav_list_ind = codici_solo_rinnovabili.index(i)

                    print("Simulazione:", lista_edifici[i], ", solo rinnovabili", sav_list_ind + 1, "/",
                          len(codici_solo_rinnovabili))

                    simulation(n_dati, dck_temp_file_path_C, dck_file_path_C, i, sav_list_ind, cod_ed_RES, cod_ed_IND,
                               cod_ed_OFF, cod_ed_SHO, carat_edif_ordered, scheda_tecnica_ordered, txt_file_path_C,
                               Gen_rin, directory_del_files_C, files_to_delete_C, txt_var)

                    np.copyto(Backup_producer, Gen_rin[0])
                else:
                    np.copyto(Gen_rin[0], Backup_producer)

    return E_abs_consumers, Delta_E_pros, Gen_rin, Autocons_pros, list_El_load_pros, list_El_gen_pros


"""                                    
                                    (4.2)  def. Calcolo Quota condivisa [kWh]                                   
"""


def q_cond(Delta_E_pros, Gen_rin, E_abs_consumers, Autocons_pros, list_El_load_pros, list_El_gen_pros, n_ed,
           codici_edifici_consumer, codici_edifici_prosumer, codici_solo_rinnovabili):
    Q_cond_tot_h = np.zeros(8760)
    # Q_cond_tot_annua = 0
    Export_inut = np.zeros(8760)
    # Export_inut_annuo = 0
    Import_cer = np.zeros(8760)
    # Import_cer_annuo = 0
    E_abs_tot_h = np.zeros(8760)
    # E_abs_tot_annua = 0
    E_gen_tot_h = np.zeros(8760)
    # E_gen_tot_annua = 0
    Autocons_h_CER = np.zeros(8760)
    # Autocons_annuo = 0
    El_load_pros_h = np.zeros(8760)
    El_load_CER = np.zeros(8760)
    # El_load_CER_annuo = 0
    El_gen_pros_h = np.zeros(8760)
    El_gen_CER = np.zeros(8760)
    # El_gen_CER_annuo = 0
    El_gen_imp_terra_annuo = 0

    for a in range(8760):
        # A: lista contenente valori di energia oraria assorbita dagli edifici
        # B: lista contenente valori di energia oraria immessa in rete dagli edifici
        # D: lista contenente valori di energia oraria autoconsumata dagli edifici
        # E: lista El_load_consumer (oraria)
        # F: lista El_gen_imp.a.terra (oraria)
        A, B, D, E, F = np.zeros(n_ed), np.zeros(n_ed), np.zeros(n_ed), np.zeros(n_ed), np.zeros(n_ed)

        for b in range(n_ed):
            if b in codici_edifici_consumer:
                indice_cons = codici_edifici_consumer.index(b)
                A[b], E[b] = E_abs_consumers[indice_cons][a], E_abs_consumers[indice_cons][a]
            elif b in codici_edifici_prosumer:
                indice_pros = codici_edifici_prosumer.index(b)
                El_load_pros_h[a] += list_El_load_pros[indice_pros][a]
                El_gen_pros_h[a] += list_El_gen_pros[indice_pros][a]
                D[b] = Autocons_pros[indice_pros][a]
                c = Delta_E_pros[indice_pros][a]
                if c < 0:
                    A[b] = -c
                elif c > 0:
                    B[b] = c
            elif b in codici_solo_rinnovabili:
                indice_producer = codici_solo_rinnovabili.index(b)
                B[b], F[b] = Gen_rin[indice_producer][a], Gen_rin[indice_producer][a]

            #
        #

        E_abs_tot_h[a] = np.sum(A)  # [kWh] energia oraria assorbita dagli edifici
        E_gen_tot_h[a] = np.sum(B)  # [kWh] energia oraria immessa in rete dagli edifici
        Q_cond_tot_h[a] = np.minimum(E_abs_tot_h[a], E_gen_tot_h[a])  # [kWh] energia virtuale condivisa
        # delta_Egen_Eabs = E_gen_tot_h[a] - E_abs_tot_h[a]  # [kWh]
        Autocons_h_CER[a] = np.sum(D)  # [kWh] energia oraria autoconsumata dagli edifici
        El_load_CER[a] = El_load_pros_h[a] + np.sum(E)  # [kWh] carico el. CER
        El_gen_CER[a] = El_gen_pros_h[a] + np.sum(F)  # [kWh]   energia generata dalla CER
        El_gen_imp_terra_annuo += np.sum(F)  # [kWh] energia generata dai producer
        #
        Export_inut[a] = np.maximum(0, (E_gen_tot_h[a] - E_abs_tot_h[a]))
        Import_cer[a] = np.maximum(0, (E_abs_tot_h[a] - E_gen_tot_h[a]))
        """
        if delta_Egen_Eabs > 0:
            Export_inut[a] = delta_Egen_Eabs
        elif delta_Egen_Eabs < 0:
            Import_cer[a] = delta_Egen_Eabs
        """
        #
    #
    E_abs_tot_annua = np.sum(E_abs_tot_h)
    E_gen_tot_annua = np.sum(E_gen_tot_h)
    Export_inut_annuo = np.sum(Export_inut)
    Import_cer_annuo = np.sum(Import_cer)
    El_load_CER_annuo = np.sum(El_load_CER)
    El_gen_CER_annuo = np.sum(El_gen_CER)
    Autocons_annuo = np.sum(Autocons_h_CER)
    Q_cond_tot_annua = np.sum(Q_cond_tot_h)

    return Q_cond_tot_annua, Q_cond_tot_h, E_abs_tot_h, E_gen_tot_h, E_abs_tot_annua, E_gen_tot_annua, Export_inut_annuo, Import_cer_annuo, Autocons_h_CER, Autocons_annuo, El_load_CER, El_load_CER_annuo, El_gen_CER, El_gen_CER_annuo, El_gen_imp_terra_annuo


"""
                    (4.3)  def. Calcolo premio CER in base alla singola Q_cond_h di ogni ed.           
"""


def premio_CER(n_ed, codici_edifici_consumer, codici_edifici_prosumer, codici_solo_rinnovabili, cod_ed_RES, cod_ed_IND,
               cod_ed_OFF, cod_ed_SHO, Q_cond_tot_h, E_abs_consumers, E_abs_tot_h, Delta_E_pros, Gen_rin, E_gen_tot_h,
               PUN_list, pot_CER):
    Premio_tot_h = np.zeros(8760)

    "Calcolo quota condivisa oraria suddivisa per ogni edificio"

    q_cond_oraria_edifici = np.zeros((n_ed, 8760))
    p_orario_edifici = np.zeros((n_ed, 8760))
    p_annuale_edifici = np.zeros(n_ed)

    for h in range(0, 8760):

        Incentivo = t_inc_eval(pot_CER, PUN_list[h])  # calcolo tariffa [€/MWh] in funzione del PUN etc.
        Premio_tot_h[h] = Q_cond_tot_h[h] * (Incentivo / 1000)  # [€]

        if E_abs_tot_h[h] != 0 and E_gen_tot_h[h] != 0:

            for ed in range(0, n_ed):

                if ed in codici_edifici_consumer:

                    i1 = codici_edifici_consumer.index(ed)

                    q_cond_oraria_edifici[ed][h] = (E_abs_consumers[i1][h] / E_abs_tot_h[h]) * Q_cond_tot_h[h]
                    p_orario_edifici[ed][h] = (q_cond_oraria_edifici[ed][h] / Q_cond_tot_h[h]) * Premio_tot_h[h] * 0.55

                elif ed in codici_edifici_prosumer:

                    i2 = codici_edifici_prosumer.index(ed)

                    if Delta_E_pros[i2][h] > 0:

                        q_cond_oraria_edifici[ed][h] = (Delta_E_pros[i2][h] / E_gen_tot_h[h]) * Q_cond_tot_h[h]
                        p_orario_edifici[ed][h] = (q_cond_oraria_edifici[ed][h] / Q_cond_tot_h[h]) * Premio_tot_h[
                            h] * 0.35

                    else:  # se si comporta come un consumer

                        q_cond_oraria_edifici[ed][h] = ((-1 * Delta_E_pros[i2][h]) / E_abs_tot_h[h]) * Q_cond_tot_h[h]
                        p_orario_edifici[ed][h] = (q_cond_oraria_edifici[ed][h] / Q_cond_tot_h[h]) * Premio_tot_h[
                            h] * 0.55

                else:

                    i3 = codici_solo_rinnovabili.index(ed)

                    q_cond_oraria_edifici[ed][h] = (Gen_rin[i3][h] / E_gen_tot_h[h]) * Q_cond_tot_h[h]
                    p_orario_edifici[ed][h] = (q_cond_oraria_edifici[ed][h] / Q_cond_tot_h[h]) * Premio_tot_h[h] * 0.35

                p_annuale_edifici[ed] += p_orario_edifici[ed][h]
            #
        #
    #
    premio_annuale_CER = np.sum(Premio_tot_h)

    "Creo una lista che contiene i premi delle utenze suddivise per tipologia e se sono pros/cons"
    premi_utenze = np.zeros(10)

    if len(cod_ed_RES) != 0:
        "RES_pros"
        premi_utenze[0] = p_annuale_edifici[cod_ed_RES[0]]
        "RES_cons"
        premi_utenze[1] = p_annuale_edifici[cod_ed_RES[len(cod_ed_RES) - 1]]
    if len(cod_ed_OFF) != 0:
        "OFF_pros"
        premi_utenze[2] = p_annuale_edifici[cod_ed_OFF[0]]
        "OFF_cons"
        premi_utenze[3] = p_annuale_edifici[cod_ed_OFF[len(cod_ed_OFF) - 1]]

    if len(cod_ed_IND) != 0:
        "IND_pros"
        premi_utenze[4] = p_annuale_edifici[cod_ed_IND[0]]
        "IND_cons"
        premi_utenze[5] = p_annuale_edifici[cod_ed_IND[len(cod_ed_IND) - 1]]

    if len(cod_ed_SHO) != 0:
        "SHO_pros"
        premi_utenze[6] = p_annuale_edifici[cod_ed_SHO[0]]
        "SHO_cons"
        premi_utenze[7] = p_annuale_edifici[cod_ed_SHO[len(cod_ed_SHO) - 1]]
    "premio impianto a terra"
    premi_utenze[8] = p_annuale_edifici[0]
    "premio totale CER"
    premi_utenze[9] = premio_annuale_CER

    return premi_utenze


"""
                                (4.4)  def. plot profili CER             
"""


def plot_CER(El_Laod_CER, El_Generation, Self_Consumption, En_Shared):
    """
    grafici: profili dei consumi,produzione,autoconsumo e quota virtuale scambiata
    """
    # Traccia le funzioni
    plt.plot(hours, El_Laod_CER, label='El_Load')
    plt.plot(hours, El_Generation, label='El_Generation')
    plt.plot(hours, Self_Consumption, label='Self-Consumption')
    plt.plot(hours, En_Shared, label='En_Shared')
    # aggiunta legenda
    plt.legend()
    # Aggiunta etichette assi e titolo
    plt.xlabel('hours')
    plt.ylabel('[MWh]')
    plt.title('Profili annuali CER')
    # Mostra il grafico
    plt.show()


"""
                                (4.5) Salvataggio risultati excel
"""


def save_ris(j, excel_data_path_save_ris, excel_data_path_save_premi, lista_save_risultati, premi_utenze):
    excel_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                     'U', 'V', 'W', 'X', 'Y', 'Z']

    """save data on Excel file"""
    # Load the Excel file into a DataFrame
    df_1 = pd.read_excel(excel_data_path_save_ris)

    for n in range(0, len(lista_save_risultati)):
        df_1.at[str(7 + j), excel_columns[n]] = lista_save_risultati[n]

    # Save the modified DataFrame to a new Excel file
    df_1.to_excel(excel_data_path_save_ris, index=False)

    df_2 = pd.read_excel(excel_data_path_save_premi)

    df_2.at[str(7 + j), excel_columns[0]] = lista_save_risultati[0]
    df_2.at[str(7 + j), excel_columns[1]] = lista_save_risultati[1]
    df_2.at[str(7 + j), excel_columns[2]] = lista_save_risultati[2]

    for n in range(3, len(premi_utenze) + 3):
        df_2.at[str(7 + j), excel_columns[n]] = premi_utenze[n - 3]
    df_2.to_excel(excel_data_path_save_premi, index=False)

    return


"""
                                        (4.6)  def. p_opt
"""


def p_opt_eval(Q_cond_tot_annua, Export_inut_annuo, Import_cer_annuo, Autocons_annuo,
               El_load_CER_annuo, El_gen_CER_annuo):
    """
                𝑤_4∙𝑝_4 + 𝑤_5∙𝑝_5
    𝑃_𝑜𝑝𝑡 = ----------------------------        (rapporto fra somme pesate)
            𝑤_1∙𝑝_1 + 𝑤_2∙𝑝_2 + 𝑤_3∙𝑝_3
    """
    "definizione pesi   NB: w_1 + w_2 + w_3 + w_4 + w_5 = 1 "
    w_1 = 0.15  # p renewable fraction
    w_2 = 0.4  # p energy sharing
    # se si da più importanza alla renewable fraction
    # possibile modifica di w1 e w2 con una if condition
    w_3 = 0.15
    w_4 = 0.15
    w_5 = 0.15

    "definizione parametri"
    p_1 = ((Q_cond_tot_annua + Autocons_annuo) / El_load_CER_annuo)  # renewable fraction
    p_2 = (Q_cond_tot_annua / El_load_CER_annuo)  # energy share
    p_3 = (El_gen_CER_annuo / El_load_CER_annuo)
    p_4 = (Import_cer_annuo / El_load_CER_annuo)
    p_5 = (Export_inut_annuo / El_gen_CER_annuo)

    "calcolo p_opt (da massimizzare)"
    # p_opt_to_sigmoid = (w_1 * p_1 + w_2 * p_2 + w_3 * p_3) / (w_4 * p_4 + w_5 * p_5)
    # p_opt = 1 / (1 + np.exp(-p_opt_to_sigmoid))

    "calcolo p_opt (da minimizzare)"
    p_opt_to_sigmoid = (w_4 * p_4 + w_5 * p_5) / (w_1 * p_1 + w_2 * p_2 + w_3 * p_3)
    p_opt = 1 / (1 + np.exp(-p_opt_to_sigmoid))

    return p_opt


"""
                                (4.7)  calcolo tariffa incentivante CER (dal sito GSE)
"""


def t_inc_eval(pot_impianto, prezzo_en):
    """
    pot_impianto = potenza installata impianto [kW]
    determina la quota fissa dell'incentivo

    prezzo_en = valore prezzo di mercato energia [€/MWh]
    determina la quota variabile dell'incentivo
    """

    "Valorizzazione energia"
    "Corrispettivo unitario"
    CU = 8.48  # [€/MWh]

    FC_zonale = 10  # per regioni del nord
    # 4 per regioni del centro

    "contributo in conto capitale"
    F = 0.05

    if pot_impianto < 200:
        A = 120  # [€/MWh]
        TP_base = 80
    elif pot_impianto < 600:
        A = 110  # [€/MWh]
        TP_base = 70
    else:
        A = 100  # [€/MWh]
        TP_base = 60
    B = TP_base + max(0, 180 - prezzo_en)

    tariffa_inc_h = (min(A, B) + FC_zonale) * (1-F) + CU  # [€/MWh]

    return tariffa_inc_h


"""
                                    Simulazione completa di una CER             
"""
"""
                                        (5.1)  def. sim. CER             
"""


def sim_CER(j, n_car, n_dati, excel_data_path_save_ris, excel_data_path_save_premi, ex_data_path_template,
            denom_p_c_ratio, ratio_RES, ratio_OFF, ratio_IND, ratio_SHO):
    """j = n°simulazione"""
    """ vettore dati da salvare su txt"""
    txt_var = [["\n", "_________________ Simulazione C.E.R. n°", j, "_________________", "\n"]]
    # Measuring time (start point)
    start_time = time.time()
    # Generazione file excel
    lista_save_ris, n_ed_tot, somma_peak_power = excel_gen(n_car, n_dati, ex_data_path_template, denom_p_c_ratio,
                                                           ratio_RES, ratio_OFF, ratio_IND, ratio_SHO)
    # import dati
    # pun
    PUN_list = collect_PUN()
    # dati CER
    [n_ed, codici_edifici_consumer, codici_edifici_prosumer, codici_solo_rinnovabili, lista_edifici,
     scheda_tecnica_ordered, carat_edif_ordered, cod_ed_RES, cod_ed_IND, cod_ed_OFF, cod_ed_SHO,
     area_PV] = collect_data(ex_data_path_template, n_car, n_dati, txt_var, n_ed_tot)

    # Avvio simulazione

    [E_abs_consumers, Delta_E_pros, Gen_rin, Autocons_pros,
     list_El_load_pros, list_El_gen_pros] = sim_edifici(n_dati, n_ed, codici_edifici_consumer, codici_edifici_prosumer,
                                                        codici_solo_rinnovabili, lista_edifici, cod_ed_RES, cod_ed_IND,
                                                        cod_ed_OFF, cod_ed_SHO, carat_edif_ordered,
                                                        scheda_tecnica_ordered,
                                                        txt_var, Liste_backup_cons, Liste_backup_pros,
                                                        Lista_backup_producer)
    # Bilanci en. e calcolo quota condivisa

    [Q_cond_tot_annua, Q_cond_tot_h, E_abs_tot_h, E_gen_tot_h, E_abs_tot_annua, E_gen_tot_annua, Export_inut_annuo,
     Import_cer_annuo, Autocons_h_CER, Autocons_annuo, El_load_CER, El_load_CER_annuo, El_gen_CER, El_gen_CER_annuo,
     El_gen_imp_terra_annuo] = q_cond(
        Delta_E_pros, Gen_rin, E_abs_consumers, Autocons_pros, list_El_load_pros, list_El_gen_pros, n_ed,
        codici_edifici_consumer, codici_edifici_prosumer, codici_solo_rinnovabili)

    # aggiornamento lista da printare su excel

    lista_save_ris.append(round(E_abs_tot_annua / 1000, 2))
    lista_save_ris.append(round(E_gen_tot_annua / 1000, 2))
    lista_save_ris.append(round(Q_cond_tot_annua / 1000, 2))
    lista_save_ris.append(round(Autocons_annuo / 1000, 2))
    lista_save_ris.append(round(El_load_CER_annuo / 1000, 2))
    lista_save_ris.append(round(El_gen_CER_annuo / 1000, 2))
    lista_save_ris.append(round(El_gen_imp_terra_annuo / 1000, 2))
    lista_save_ris.append(round(Import_cer_annuo / 1000, 2))
    lista_save_ris.append(round(Export_inut_annuo / 1000, 2))
    lista_save_ris.append(round(((Q_cond_tot_annua + Autocons_annuo) / El_load_CER_annuo), 2))
    lista_save_ris.append(round((Q_cond_tot_annua / El_gen_CER_annuo), 2))
    lista_save_ris.append(round((Q_cond_tot_annua / El_load_CER_annuo), 2))
    lista_save_ris.append(round((El_gen_CER_annuo / El_load_CER_annuo), 2))
    lista_save_ris.append(round((Export_inut_annuo / El_gen_CER_annuo), 2))
    lista_save_ris.append(round((Import_cer_annuo / El_load_CER_annuo), 2))

    # Calcolo p_opt

    p_opt = p_opt_eval(Q_cond_tot_annua, Export_inut_annuo, Import_cer_annuo, Autocons_annuo,
                       El_load_CER_annuo, El_gen_CER_annuo)

    lista_save_ris.append(round(p_opt, 2))

    # Plot profili carico,gen,autoconsumo,q_cond

    #plot_CER(El_load_CER, El_gen_CER, Autocons_h_CER, Q_cond_tot_h)

    # Calcolo premi CER

    premi_utenze = premio_CER(n_ed, codici_edifici_consumer, codici_edifici_prosumer, codici_solo_rinnovabili,
                              cod_ed_RES, cod_ed_IND, cod_ed_OFF, cod_ed_SHO, Q_cond_tot_h, E_abs_consumers,
                              E_abs_tot_h, Delta_E_pros, Gen_rin, E_gen_tot_h, PUN_list, somma_peak_power)

    # salvataggio dati excel

    save_ris(j, excel_data_path_save_ris, excel_data_path_save_premi, lista_save_ris, premi_utenze)

    # Calcolo tempo impiegato per la simulazione CER

    elapsed_time = time.time() - start_time  # Measuring time (end point)
    print("")
    print("elapsed time for complete simulation:", round(elapsed_time, 2), "[s]")

    return p_opt


"""
                                (5.2) Input simulazione CER
"""

"   Input manuali   "

save_txt_file_path = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Output_Risultati_CER.txt"

"in base al file excel [non modificare]"

n_car = 7
n_dati = 10

"Files excel"

ex_data_path_template = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Foglio raccolta dati C.E.R._TEMPLATE.xlsx"

save_excel_path = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Foglio_risultati_simulazione_TEMPLATE.xlsx"

excel_data_path_save_premi = r"C:\Users\Mirco Tirloni\Desktop\TESI_Mirco\Codice_TESI_Mirco\Foglio_premi_utenze_TEMPLATE.xlsx"

"contatore numero simulazioni"

n_sim = 0
